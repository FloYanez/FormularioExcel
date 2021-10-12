import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook


class Register(tk.Tk):
    def __init__(self):
        super().__init__()

        self.geometry('400x550')  # 300x150
        self.resizable(0, 0)
        self.title('Registro de Datos')
        self.iconbitmap('icon.ico')

        # UI options
        paddings = {'padx': 5, 'pady': 5}
        entry_font = {'font': ('Helvetica', 11)}

        # configure the grid
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=3)

        # heading
        heading = ttk.Label(self, text='Registro de Datos de Operaciones', style='Heading.TLabel')
        heading.grid(column=0, row=0, columnspan=2, pady=5, sticky=tk.N)

        # Fecha
        fecha = tk.StringVar()
        fecha_label = ttk.Label(self, text="Fecha:")
        fecha_label.grid(column=0, row=1, sticky=tk.W, **paddings)
        self.fecha_entry = ttk.Entry(self, textvariable=fecha, **entry_font)
        self.fecha_entry.grid(column=1, row=1, sticky=tk.E, **paddings)

        # Número de operación
        num_op = tk.IntVar()
        num_op_label = ttk.Label(self, text="N° de Operación:")
        num_op_label.grid(column=0, row=2, sticky=tk.W, **paddings)
        self.num_op_entry = ttk.Entry(self, textvariable=num_op, **entry_font)
        self.num_op_entry.grid(column=1, row=2, sticky=tk.E, **paddings)

        # Código
        codigo = tk.IntVar()
        codigo_label = ttk.Label(self, text="Código:")
        codigo_label.grid(column=0, row=3, sticky=tk.W, **paddings)
        self.codigo_entry = ttk.Entry(self, textvariable=codigo, **entry_font)
        self.codigo_entry.grid(column=1, row=3, sticky=tk.E, **paddings)

        # Cantidad programada
        programado = tk.IntVar()
        programado_label = ttk.Label(self, text="Programado:")
        programado_label.grid(column=0, row=4, sticky=tk.W, **paddings)
        self.programado_entry = ttk.Entry(self, textvariable=programado, **entry_font)
        self.programado_entry.grid(column=1, row=4, sticky=tk.E, **paddings)

        # Efectivo
        efectivo = tk.IntVar()
        efectivo_label = ttk.Label(self, text="Efectivo:")
        efectivo_label.grid(column=0, row=5, sticky=tk.W, **paddings)
        self.efectivo_entry = ttk.Entry(self, textvariable=efectivo, **entry_font)
        self.efectivo_entry.grid(column=1, row=5, sticky=tk.E, **paddings)

        # Defectuosos
        defectuosos = tk.IntVar()
        defectuosos_label = ttk.Label(self, text="Defectuosos:")
        defectuosos_label.grid(column=0, row=6, sticky=tk.W, **paddings)
        self.defectuosos_entry = ttk.Entry(self, textvariable=defectuosos, **entry_font)
        self.defectuosos_entry.grid(column=1, row=6, sticky=tk.E, **paddings)

        # Hora Inicio OP
        inicio = tk.StringVar()
        inicio_label = ttk.Label(self, text="Hora Inicio OP:")
        inicio_label.grid(column=0, row=7, sticky=tk.W, **paddings)
        self.inicio_entry = ttk.Entry(self, textvariable=inicio, **entry_font)
        self.inicio_entry.grid(column=1, row=7, sticky=tk.E, **paddings)

        # Hora Término OP
        termino = tk.StringVar()
        termino_label = ttk.Label(self, text="Hora Término OP:")
        termino_label.grid(column=0, row=8, sticky=tk.W, **paddings)
        self.termino_entry = ttk.Entry(self, textvariable=termino, **entry_font)
        self.termino_entry.grid(column=1, row=8, sticky=tk.E, **paddings)

        # Dotación
        dotacion = tk.IntVar()
        dotacion_label = ttk.Label(self, text="Dotación:")
        dotacion_label.grid(column=0, row=9, sticky=tk.W, **paddings)
        self.dotacion_entry = ttk.Entry(self, textvariable=dotacion, **entry_font)
        self.dotacion_entry.grid(column=1, row=9, sticky=tk.E, **paddings)

        # HHEE
        horas_extras = tk.IntVar()
        horas_extras_label = ttk.Label(self, text="Horas Extra:")
        horas_extras_label.grid(column=0, row=10, sticky=tk.W, **paddings)
        self.horas_extras_entry = ttk.Entry(self, textvariable=horas_extras, **entry_font)
        self.horas_extras_entry.grid(column=1, row=10, sticky=tk.E, **paddings)

        # Paros Programados
        paros_programados = tk.DoubleVar()
        paros_programados_label = ttk.Label(self, text="Paros Programados:")
        paros_programados_label.grid(column=0, row=11, sticky=tk.W, **paddings)
        self.paros_programados_entry = ttk.Entry(self, textvariable=paros_programados, **entry_font)
        self.paros_programados_entry.grid(column=1, row=11, sticky=tk.E, **paddings)

        # Paros por Fallas
        paros_por_fallas = tk.DoubleVar()
        paros_por_fallas_label = ttk.Label(self, text="Paros por Fallas:")
        paros_por_fallas_label.grid(column=0, row=12, sticky=tk.W, **paddings)
        self.paros_por_fallas_entry = ttk.Entry(self, textvariable=paros_por_fallas, **entry_font)
        self.paros_por_fallas_entry.grid(column=1, row=12, sticky=tk.E, **paddings)

        # Descripción PP
        descripcion_pp = tk.StringVar()
        descripcion_pp_label = ttk.Label(self, text="Descripción Paros Programados:")
        descripcion_pp_label.grid(column=0, row=13, sticky=tk.W, **paddings)
        self.descripcion_pp_entry = ttk.Entry(self, textvariable=descripcion_pp, **entry_font)
        self.descripcion_pp_entry.grid(column=1, row=13, sticky=tk.E, **paddings)

        # Descripción PnP
        descripcion_ppf = tk.StringVar()
        descripcion_ppf_label = ttk.Label(self, text="Descripción Paros por Fallas:")
        descripcion_ppf_label.grid(column=0, row=14, sticky=tk.W, **paddings)
        self.descripcion_ppf_entry = ttk.Entry(self, textvariable=descripcion_ppf, **entry_font)
        self.descripcion_ppf_entry.grid(column=1, row=14, sticky=tk.E, **paddings)

        # login button
        guardar = ttk.Button(self, text="Guardar", command=self.guardar)
        guardar.grid(column=1, row=15, sticky=tk.E, **paddings)

        # configure style
        self.style = ttk.Style(self)
        self.style.configure('TLabel', font=('Helvetica', 11))
        self.style.configure('TButton', font=('Helvetica', 11))

        # heading style
        self.style.configure('Heading.TLabel', font=('Helvetica', 12))

    def collect_data(self):
        return [self.fecha_entry.get(),
                self.num_op_entry.get(),
                self.codigo_entry.get(),
                self.programado_entry.get(),
                self.efectivo_entry.get(),
                self.defectuosos_entry.get(),
                self.inicio_entry.get(),
                self.termino_entry.get(),
                self.dotacion_entry.get(),
                self.horas_extras_entry.get(),
                self.paros_programados_entry.get(),
                self.paros_por_fallas_entry.get(),
                self.descripcion_pp_entry.get(),
                self.descripcion_ppf_entry.get()]
    
    def guardar(self):
        data = self.collect_data()
        file_name = "Registro Datos.xlsx"
        wb = load_workbook(file_name)
        ws = wb["BD"]
        row = ws.max_row + 1

        for column in range(2, 16):
            ws.cell(row=row, column=column).value = data[column - 2]

        wb.template = False
        wb.save(file_name)


if __name__ == "__main__":
    app = Register()
    app.mainloop()
